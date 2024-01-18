#-*-coding: utf-8-*-

import os
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
import datetime
from openpyxl.chart import LineChart, Reference, BarChart

# 최근 비번 변경일자 ~ 오늘까지 90일 넘었는지 체크
def timecalc(intext):
  today = datetime.datetime.now()
  sep = intext.index('-')
  year = int(intext[sep - 4:sep])
  month = int(intext[sep + 1:sep + 3])
  day = int(intext[sep + 4:sep + 6])
  return (today - (datetime.datetime(year, month, day))).days

# 패스워드 정책 설정 일 수 뽑는 함수
def passrule(intext):
  sep = intext.index('=')
  return int(intext[sep + 2:])

# path = "파일을 끌어올 경로"
wb = openpyxl.load_workbook(path)
# 맨 처음 시트
default = wb.active
# 결과 뽑을 result 시트 만들기
ws = wb.create_sheet('result')
sheet = wb['result']
# 그래프 저장할 pivot 시트 만들기
pivot = wb.create_sheet('result')
pivot_sheet = wb['pivot']

# 1행에 범례 추가
sheet.cell(row = 1, column = 1, value = "사번")
sheet.cell(row = 1, column = 2, value = "이름")
sheet.cell(row = 1, column = 3, value = "부서")
sheet.cell(row = 1, column = 4, value = "항목1")
sheet.cell(row = 1, column = 5, value = "항목2")
sheet.cell(row = 1, column = 6, value = "항목3")
sheet.cell(row = 1, column = 7, value = "항목4")
sheet.cell(row = 1, column = 8, value = "항목5")

# path에서 파일 목록 출력(마지막 하나는 폴더라서 제외)
flist = os.listdir(path)[:-1]

# 한 파일이 끝나면 다음 열로 넘어가기 위한 cnt 선언
# 1행은 범례, 2행부터 실제 데이터가 들어가야 하므로 2부터 시작
# cnt는 카운트 외에 데이터를 몇 행부터 쓸건지에 사용
cnt = 2

# 취약한 설정 보유한 사용자 리스트
vul_user = []

for i in flist:
  # 취약한 내용 저장
  content = []
  # 사용자 PC 점검 결과 저장 리스트
  tmp = []
  # 총점
  vscore = 0

  #취약한 부분 판별 위한 항목당 점수 선언
  share = 0
  last = 0
  minage = 0
  maxage = 0
  minlen = 0
  complx = 0
  hisize = 0
  clspass = 0

  
  f = open(path+i, 'rb')
  
  for j in f:
    j = j.decode(encoding='unicode-escape')
    # 쓸데없는 공백 지우기 위함
    j = j.strip()
    j = str(j)
    # tmp라는 리스트에 파일의 내용을 한줄씩 입력
    tmp.append(j)

  # 취약 여부 판단하는 부분
  # pc에서 autotest.bat를 돌려서 나온 result.txt를 tmp에 담아 키워드로 취약을 판단
  for ck in tmp:
    if 'C$' in ck or 'D$' in ck or 'E$' in ck or 'F$' in ck or 'ADMIN$' in ck:
      share += 1
    if 'Password last set' in ck:
      if timecalc(ck) > 90:
        last += 1
        content.append('Password last set over')
      if 'MinimumPasswordAge' in ck:
        if passrule(ck) == 0:
          minage += 1
          content.append('Minimum Password Age is too short')
      if 'MaximumPasswordAge' in ck:
        if passrule(ck) < 90:
          maxage += 1
          content.append('Maximum Password Age is too big')
      if 'MinimumPasswordLength' in ck:
        if passrule(ck) < 8:
          minlen += 1
          content.append('Minumum Password Length is too short')
      if 'PasswordComplexity' in ck:
        if passrule(ck) == 0:
          complx += 1
          content.append('Password Complexity is not apply')
        if 'PasswordHistorySize' in ck:
          if passrule(ck) < 5:
            hisize += 1
            content.append('Password History Size is too small')
        if 'ClearTextPassword' in ck:
          if passrule(ck) == 1:
            clspass += 1
            content.append('Cleartext can\'t be password')

  # 취약하지 않은 사용자는 넘어감
  if len(content) == 0:
    pass
  # 취약한 사용자만 엑셀에 사번, 이름, 취약한 항목 기록
  else:
    # 합산
    vscore = share + last + minage + maxage + minlen + complx + hisize + clspass
    if vscore > 1:
      vul_user.append(i)
      # 취약한 경우 사번과 사유를 엑셀에 적음
      sheet.cell(row = cnt, column = 1, value = ILLGEAL_CHARACTERS_RE.sub(r'',i[:9]))
      # 취약 사유가 담긴 content를 D열부터 한개씩 넣음.. content=['a','b','c']인 경우 D열 = a, E열 = b, F열 = c 들어감
      for rea in range(len(content)):
        sheet.cell(row = cnt, column = 4 + rea, value = ILLEGAL_CHARACTERS_RE.sub(r'', content[rea]))
    # result.txt 닫음
    f.close()

    #sheet0의 A열 개수 카운팅 위한 변수
    arow = 0
    for r in default.rows:
      arow += 1
    # IAM에 들어있는 전사 직원들의 사번 갯수
    arow = str(arow)

    # 부서, sheet0에서 사번을 기준으로 vlookup하여 이름 갖고옴
    sheet.cell(row = cnt, column = 2, value = "=VLOOKUP(result!A" + str(cnt) + ",sheet0!A1:D" + arow + ",2,0)")
    # 부서, sheet0에서 사번을 기준으로 vlookup하여 부서 갖고옴
    sheet.cell(row = cnt, column = 3, value = "=VLOOKUP(result!A" + str(cnt) + ",sheet0!A1:D" + arow + ",4,0)")
    cnt += 1

# result 시트의 부서를 카운트 하기 위한 부분
# C열의 부서 목록에서 중복을 제거한 부서 목록을 N열에 놓고
# C열에 있는 부서들의 갯수를 카운팅해서 O열에 놓음
team = ["A팀", "B팀", "C팀" ....]

# N열에 부서, O열에 검출수 범례 추가
sheet.cell(row=1, column=14, value="부서")
sheet.cell(row=1, column=15, value="검출수")

# 부서당 검출수 뽑기 위해 전체 부서 목록인 team 리스트를 N열 2행부터 집어넣음
for qq in range(0, len(team)):
  sheet.cell(row=qq + 2, column = 14, value = team[qq])
#부서당 검출수 뽑기 위해 countif 사용
for ww in ragne(0, len(team)):
  sheet.cell(row=ww + 2, column = 15, value = "=COUNTIF($C$2:$C$" + str(cnt - 1) + ",N" + str(ww + 2) + ")")

# pivot이라는 시트의 B2에 차트 추가함
pivot_table = pivot_sheet['B2']

chart = BarChart()
chart.title = "검출 부서 수"
chart.style = 13
chart.y_axis.title = "검출수"
chart.x_axis.title = "부서명"

# 부서당 검출 수 
values = Reference(sheet, min_col=15, max_col=15, min_row=2, max_row=len(team) + 1)
# 부서 목록
cat1 = Reference(sheet, min_col=14, max_col=14, min_row=2, max_row=len(team) + 1)
chart.add_data(values)
chart.set_catagories(cat1)
pivot.add_chart(chart, "B2")

wb.save("path")
